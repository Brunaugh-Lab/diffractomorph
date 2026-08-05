"""Read BioTek plate exports — the Time|280|490 kinetic table."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl


@dataclass
class PlateTimecourse:
    times_min: np.ndarray
    a280: np.ndarray
    a490: np.ndarray
    blank280: float          # per-plate blank; NaN if the plate has no Blank row
    blank490: float
    source: str = ""


@dataclass
class WavelengthPlateTimecourse:
    """Kinetic plate trajectory for a caller-declared wavelength set."""

    times_min: np.ndarray
    absorbance: dict[int, np.ndarray]
    blank: dict[int, float]
    source: str = ""


def read_plate_wavelengths(path, wavelengths_nm) -> WavelengthPlateTimecourse:
    """Parse ``Time|<wavelengths...>`` using only caller-declared wavelengths."""
    path = Path(path)
    wavelengths = tuple(int(value) for value in wavelengths_nm)
    if not wavelengths or len(set(wavelengths)) != len(wavelengths):
        raise ValueError("wavelengths_nm must be a non-empty unique sequence")
    rows = list(openpyxl.load_workbook(path, data_only=True).worksheets[0].iter_rows(values_only=True))
    hr = hc = None
    for i, row in enumerate(rows):
        for j, value in enumerate(row):
            if not (isinstance(value, str) and value.strip().lower() == "time"):
                continue
            observed = tuple(int(v) for v in row[j + 1:j + 1 + len(wavelengths)]
                             if isinstance(v, (int, float)) and not isinstance(v, bool))
            if observed == wavelengths:
                hr, hc = i, j
                break
        if hr is not None:
            break
    if hr is None:
        labels = "|".join(str(value) for value in wavelengths)
        raise ValueError(f"{path.name}: no Time|{labels} kinetic header found")
    times: list[float] = []
    absorbance = {wavelength: [] for wavelength in wavelengths}
    blanks = {wavelength: np.nan for wavelength in wavelengths}
    for row in rows[hr + 1:]:
        time_value = row[hc]
        values = row[hc + 1:hc + 1 + len(wavelengths)]
        if isinstance(time_value, str) and "blank" in time_value.lower():
            for wavelength, value in zip(wavelengths, values):
                blanks[wavelength] = float(value)
        elif isinstance(time_value, (int, float)) and not isinstance(time_value, bool):
            if any(value is None for value in values):
                raise ValueError(f"{path.name}: incomplete kinetic row at time {time_value}")
            times.append(float(time_value))
            for wavelength, value in zip(wavelengths, values):
                absorbance[wavelength].append(float(value))
        elif time_value is None and times:
            break
    if not times:
        raise ValueError(f"{path.name}: no kinetic rows under requested header")
    return WavelengthPlateTimecourse(
        np.asarray(times, float),
        {wavelength: np.asarray(values, float) for wavelength, values in absorbance.items()},
        {wavelength: float(value) for wavelength, value in blanks.items()},
        path.name,
    )


def read_plate(path) -> PlateTimecourse:
    """Parse a plate export → the Time|280|490 trajectory + per-plate blank."""
    generic = read_plate_wavelengths(path, (280, 490))
    return PlateTimecourse(
        generic.times_min, generic.absorbance[280], generic.absorbance[490],
        generic.blank[280], generic.blank[490], generic.source,
    )


@dataclass
class QCRead:
    """One endpoint suspension read from a plate grid: replicate wells + blank."""
    wells: list          # populated well labels, e.g. ["E5", "F5", "G5"]
    values: list         # raw replicate absorbances (blank excluded)
    blank: float         # the block's blank (or the file's, if this block has none)
    abs_bs: float        # blank-subtracted mean of ``values``
    wavelength_nm: int


def _col_header(r):
    """Column index where a plate grid's ``1 2 3 4 5 …`` header starts, else None."""
    for j in range(len(r) - 4):
        if all(isinstance(v, (int, float)) and not isinstance(v, bool) and v == k + 1
               for k, v in enumerate(r[j:j + 5])):
            return j
    return None


def _block_wavelength(rows, i, j):
    """The wavelength tag sitting to the right of the grid at header row ``i``, col ``j``.

    Handles both a bare numeric tag (``453``) and a labelled string (``"Read 2:453"``).
    """
    for k in range(i + 1, min(i + 9, len(rows))):
        for c in range(j + 12, min(j + 20, len(rows[k]))):
            v = rows[k][c]
            if v in (280, 453, 490, 280.0, 453.0, 490.0):
                return int(v)
            if isinstance(v, str) and (m := re.search(r"\b(280|453|490)\b", v)):
                return int(m.group(1))
    return None


def read_qc(path, wavelength=453, blank_max=0.2):
    """Find every plate grid at ``wavelength`` and return the suspension read(s).

    Locates each grid by its ``1..12`` column header + ``A–H`` row labels + wavelength tag
    (robust to where the block sits), splits each block's populated wells into a replicate
    cluster and a blank (any value below ``blank_max``), and returns one :class:`QCRead` per
    block. A block with no internal blank borrows the file's blank. Usually one read per file;
    a file with several (e.g. multiple suspensions) returns them in sheet order.
    """
    path = Path(path)
    rows = list(openpyxl.load_workbook(path, data_only=True).worksheets[0].iter_rows(values_only=True))
    blocks = []                                   # (wells, values, wavelength)
    for i, r in enumerate(rows):
        j = _col_header(r)
        if j is None:
            continue
        wl = _block_wavelength(rows, i, j)
        if wl != wavelength:
            continue
        wells, vals = [], []
        for k in range(i + 1, min(i + 9, len(rows))):
            rr = rows[k]
            label = rr[j - 1] if j > 0 else ""
            for c in range(j, min(j + 12, len(rr))):
                v = rr[c]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    wells.append(f"{label}{c - j + 1}"); vals.append(float(v))
        if vals:
            blocks.append((wells, vals, wl))
    file_blank = next((float(np.mean([v for v in vals if v < blank_max]))
                       for _, vals, _ in blocks if any(v < blank_max for v in vals)), None)
    reads = []
    for wells, vals, wl in blocks:
        lows = [v for v in vals if v < blank_max]
        highs = [(w, v) for w, v in zip(wells, vals) if v >= blank_max]
        blank = float(np.mean(lows)) if lows else file_blank
        if not highs or blank is None:
            continue
        sv = [v for _, v in highs]
        reads.append(QCRead([w for w, _ in highs], sv, blank, float(np.mean(sv)) - blank, wl))
    return reads
